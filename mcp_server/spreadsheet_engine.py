from __future__ import annotations

import os
import re
from pathlib import Path

import openpyxl
from odf import table, text
from odf.opendocument import load as load_ods

from models import (
    GetSheetInfoResponse,
    ReadCellResponse,
    ReadRangeResponse,
    SheetInfo,
    SpreadsheetInfo,
    WriteCellResponse,
    WriteRangeResponse,
)


def _col_letter_to_index(col: str) -> int:
    """Convert column letter(s) to 0-based index: A->0, B->1, ..., Z->25, AA->26."""
    result = 0
    for char in col.upper():
        result = result * 26 + (ord(char) - ord("A") + 1)
    return result - 1


def _index_to_col_letter(index: int) -> str:
    """Convert 0-based index to column letter(s): 0->A, 1->B, ..., 25->Z, 26->AA."""
    result = ""
    index += 1
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        result = chr(ord("A") + remainder) + result
    return result


def _parse_cell_ref(cell_ref: str) -> tuple[int, int]:
    """Parse 'A1' style reference into (row_index, col_index), both 0-based."""
    match = re.match(r"^([A-Za-z]+)(\d+)$", cell_ref.strip())
    if not match:
        raise ValueError(f"Invalid cell reference: '{cell_ref}'. Expected format like 'A1', 'B5', 'AA10'.")
    col_str, row_str = match.groups()
    row_num = int(row_str)
    if row_num < 1:
        raise ValueError(f"Invalid row number in cell reference: '{cell_ref}'. Rows start at 1.")
    return row_num - 1, _col_letter_to_index(col_str)


def _resolve_path(file_path: str) -> Path:
    path = Path(file_path)
    if not path.is_absolute():
        workspace = Path(os.getenv("WORKSPACE_DIR", "/workspace"))
        candidate = (workspace / path).resolve()
        if candidate.exists():
            return candidate
        cwd_candidate = Path.cwd() / path
        if cwd_candidate.resolve().exists():
            return cwd_candidate.resolve()
        raise FileNotFoundError(f"File not found: tried {candidate} and {cwd_candidate.resolve()}")
    path = path.resolve()
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")
    return path


def _get_file_type(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in (".ods",):
        return "ods"
    if suffix in (".xlsx", ".xls"):
        return "xlsx"
    if suffix in (".csv",):
        return "csv"
    raise ValueError(f"Unsupported file type: {suffix}")


def _find_ods_sheet(doc: object, sheet_name: str) -> object:
    """Look up a named sheet in an ODS document, raising ValueError if absent."""
    sheets = doc.spreadsheet.getElementsByType(table.Table)
    for s in sheets:
        if s.getAttribute("name") == sheet_name:
            return s
    raise ValueError(f"Sheet '{sheet_name}' not found")


def _read_ods_cell_value(cell: object) -> tuple[object, str]:
    """Extract (value, value_type) from an ODS TableCell element."""
    val_type = cell.getAttribute("valuetype") or "string"
    if val_type in ("float", "percentage", "currency"):
        raw = cell.getAttribute("value")
        value = float(raw) if raw else None
    else:
        paragraphs = cell.getElementsByType(text.P)
        raw_text = "".join(
            str(node) for p in paragraphs for node in p.childNodes
        ) if paragraphs else None
        value = raw_text if raw_text else None
    if value is None:
        val_type = "empty"
    return value, val_type


class OdsEngine:
    """Engine for reading/writing ODS spreadsheets using odfpy."""

    @staticmethod
    def read_cell(path: Path, sheet_name: str, cell_ref: str) -> ReadCellResponse:
        row_idx, col_idx = _parse_cell_ref(cell_ref)
        doc = load_ods(str(path))
        sheet = _find_ods_sheet(doc, sheet_name)

        rows = sheet.getElementsByType(table.TableRow)
        if row_idx >= len(rows):
            return ReadCellResponse(cell_reference=cell_ref, value=None, value_type="empty")

        cells = rows[row_idx].getElementsByType(table.TableCell)
        if col_idx >= len(cells):
            return ReadCellResponse(cell_reference=cell_ref, value=None, value_type="empty")

        value, val_type = _read_ods_cell_value(cells[col_idx])
        return ReadCellResponse(cell_reference=cell_ref, value=value, value_type=val_type)

    @staticmethod
    def read_range(path: Path, sheet_name: str, start: str, end: str) -> ReadRangeResponse:
        r1, c1 = _parse_cell_ref(start)
        r2, c2 = _parse_cell_ref(end)
        doc = load_ods(str(path))
        sheet = _find_ods_sheet(doc, sheet_name)

        rows = sheet.getElementsByType(table.TableRow)
        result: list[list[str | int | float | bool | None]] = []
        for ri in range(r1, r2 + 1):
            row_vals: list[str | int | float | bool | None] = []
            if ri < len(rows):
                cells = rows[ri].getElementsByType(table.TableCell)
                for ci in range(c1, c2 + 1):
                    if ci < len(cells):
                        value, _ = _read_ods_cell_value(cells[ci])
                        row_vals.append(value)
                    else:
                        row_vals.append(None)
            else:
                row_vals = [None] * (c2 - c1 + 1)
            result.append(row_vals)

        return ReadRangeResponse(start_cell=start, end_cell=end, values=result)

    @staticmethod
    def write_cell(
        path: Path, sheet_name: str, cell_ref: str, value: str | int | float | bool
    ) -> WriteCellResponse:
        doc = load_ods(str(path))
        sheet = _find_ods_sheet(doc, sheet_name)

        row_idx, col_idx = _parse_cell_ref(cell_ref)
        rows = sheet.getElementsByType(table.TableRow)

        while len(rows) <= row_idx:
            new_row = table.TableRow()
            sheet.addElement(new_row)
            rows = sheet.getElementsByType(table.TableRow)

        row = rows[row_idx]
        cells = row.getElementsByType(table.TableCell)
        while len(cells) <= col_idx:
            new_cell = table.TableCell()
            row.addElement(new_cell)
            cells = row.getElementsByType(table.TableCell)

        cell = cells[col_idx]
        for child in list(cell.childNodes):
            cell.removeChild(child)

        if isinstance(value, bool):
            cell.setAttribute("valuetype", "boolean")
            cell.setAttribute("booleanvalue", str(value).lower())
            p = text.P(text=str(value))
            cell.addElement(p)
        elif isinstance(value, (int, float)):
            cell.setAttribute("valuetype", "float")
            cell.setAttribute("value", str(value))
            p = text.P(text=str(value))
            cell.addElement(p)
        else:
            cell.setAttribute("valuetype", "string")
            p = text.P(text=str(value))
            cell.addElement(p)

        doc.save(str(path))
        return WriteCellResponse(success=True, cell_reference=cell_ref, new_value=value)

    @staticmethod
    def write_range(
        path: Path,
        sheet_name: str,
        start: str,
        values: list[list[str | int | float | bool | None]],
    ) -> WriteRangeResponse:
        doc = load_ods(str(path))
        sheet = _find_ods_sheet(doc, sheet_name)

        r1, c1 = _parse_cell_ref(start)
        cells_written = 0

        rows = sheet.getElementsByType(table.TableRow)
        for ri, row_vals in enumerate(values):
            actual_row = r1 + ri
            while len(rows) <= actual_row:
                sheet.addElement(table.TableRow())
                rows = sheet.getElementsByType(table.TableRow)

            row = rows[actual_row]
            cells = row.getElementsByType(table.TableCell)

            for ci, val in enumerate(row_vals):
                actual_col = c1 + ci
                while len(cells) <= actual_col:
                    row.addElement(table.TableCell())
                    cells = row.getElementsByType(table.TableCell)

                cell = cells[actual_col]
                for child in list(cell.childNodes):
                    cell.removeChild(child)

                if val is None:
                    continue

                if isinstance(val, bool):
                    cell.setAttribute("valuetype", "boolean")
                    cell.setAttribute("booleanvalue", str(val).lower())
                    cell.addElement(text.P(text=str(val)))
                elif isinstance(val, (int, float)):
                    cell.setAttribute("valuetype", "float")
                    cell.setAttribute("value", str(val))
                    cell.addElement(text.P(text=str(val)))
                else:
                    cell.setAttribute("valuetype", "string")
                    cell.addElement(text.P(text=str(val)))
                cells_written += 1

        doc.save(str(path))
        return WriteRangeResponse(success=True, cells_written=cells_written)

    @staticmethod
    def set_formula(path: Path, sheet_name: str, cell_ref: str, formula: str) -> None:
        doc = load_ods(str(path))
        sheet = _find_ods_sheet(doc, sheet_name)

        row_idx, col_idx = _parse_cell_ref(cell_ref)
        rows = sheet.getElementsByType(table.TableRow)
        while len(rows) <= row_idx:
            sheet.addElement(table.TableRow())
            rows = sheet.getElementsByType(table.TableRow)

        row = rows[row_idx]
        cells = row.getElementsByType(table.TableCell)
        while len(cells) <= col_idx:
            row.addElement(table.TableCell())
            cells = row.getElementsByType(table.TableCell)

        cell = cells[col_idx]
        cell.setAttribute("formula", formula)
        doc.save(str(path))

    @staticmethod
    def read_formula(path: Path, sheet_name: str, cell_ref: str) -> str | None:
        """Read the raw formula string from a cell, or None if no formula is set."""
        row_idx, col_idx = _parse_cell_ref(cell_ref)
        doc = load_ods(str(path))
        sheet = _find_ods_sheet(doc, sheet_name)
        rows = sheet.getElementsByType(table.TableRow)
        if row_idx >= len(rows):
            return None
        cells = rows[row_idx].getElementsByType(table.TableCell)
        if col_idx >= len(cells):
            return None
        return cells[col_idx].getAttribute("formula") or None

    @staticmethod
    def get_sheet_info(path: Path, sheet_name: str) -> GetSheetInfoResponse:
        doc = load_ods(str(path))
        sheet = _find_ods_sheet(doc, sheet_name)

        rows = sheet.getElementsByType(table.TableRow)
        row_count = len(rows)
        col_count = 0
        headers: list[str | None] = []

        if rows:
            first_row_cells = rows[0].getElementsByType(table.TableCell)
            col_count = len(first_row_cells)
            for cell in first_row_cells:
                paragraphs = cell.getElementsByType(text.P)
                val = "".join(
                    str(n) for p in paragraphs for n in p.childNodes
                ) if paragraphs else None
                headers.append(val)

        for row in rows:
            cells = row.getElementsByType(table.TableCell)
            col_count = max(col_count, len(cells))

        return GetSheetInfoResponse(
            sheet_name=sheet_name,
            row_count=row_count,
            col_count=col_count,
            headers=headers,
        )

    @staticmethod
    def list_sheets(path: Path) -> SpreadsheetInfo:
        doc = load_ods(str(path))
        sheets = doc.spreadsheet.getElementsByType(table.Table)
        sheet_infos = []
        for s in sheets:
            name = s.getAttribute("name")
            rows = s.getElementsByType(table.TableRow)
            row_count = len(rows)
            col_count = 0
            for row in rows:
                cells = row.getElementsByType(table.TableCell)
                col_count = max(col_count, len(cells))
            sheet_infos.append(SheetInfo(name=name, row_count=row_count, col_count=col_count))
        return SpreadsheetInfo(file_path=str(path), sheets=sheet_infos)


class XlsxEngine:
    """Engine for reading/writing Excel spreadsheets using openpyxl."""

    @staticmethod
    def read_cell(path: Path, sheet_name: str, cell_ref: str) -> ReadCellResponse:
        wb = openpyxl.load_workbook(str(path), data_only=True)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found")
        ws = wb[sheet_name]
        cell = ws[cell_ref]
        value = cell.value
        if value is None:
            vtype = "empty"
        elif isinstance(value, bool):
            vtype = "boolean"
        elif isinstance(value, (int, float)):
            vtype = "float"
        else:
            vtype = "string"
        return ReadCellResponse(cell_reference=cell_ref, value=value, value_type=vtype)

    @staticmethod
    def read_range(path: Path, sheet_name: str, start: str, end: str) -> ReadRangeResponse:
        wb = openpyxl.load_workbook(str(path), data_only=True)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found")
        ws = wb[sheet_name]
        result: list[list[str | int | float | bool | None]] = []
        for row in ws[f"{start}:{end}"]:
            result.append([cell.value for cell in row])
        return ReadRangeResponse(start_cell=start, end_cell=end, values=result)

    @staticmethod
    def write_cell(
        path: Path, sheet_name: str, cell_ref: str, value: str | int | float | bool
    ) -> WriteCellResponse:
        wb = openpyxl.load_workbook(str(path))
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found")
        ws = wb[sheet_name]
        ws[cell_ref] = value
        wb.save(str(path))
        return WriteCellResponse(success=True, cell_reference=cell_ref, new_value=value)

    @staticmethod
    def write_range(
        path: Path,
        sheet_name: str,
        start: str,
        values: list[list[str | int | float | bool | None]],
    ) -> WriteRangeResponse:
        wb = openpyxl.load_workbook(str(path))
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found")
        ws = wb[sheet_name]

        r1, c1 = _parse_cell_ref(start)
        cells_written = 0
        for ri, row_vals in enumerate(values):
            for ci, val in enumerate(row_vals):
                if val is not None:
                    ws.cell(row=r1 + ri + 1, column=c1 + ci + 1, value=val)
                    cells_written += 1
        wb.save(str(path))
        return WriteRangeResponse(success=True, cells_written=cells_written)

    @staticmethod
    def set_formula(path: Path, sheet_name: str, cell_ref: str, formula: str) -> None:
        wb = openpyxl.load_workbook(str(path))
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found")
        ws = wb[sheet_name]
        ws[cell_ref] = formula
        wb.save(str(path))

    @staticmethod
    def read_formula(path: Path, sheet_name: str, cell_ref: str) -> str | None:
        """Read the raw formula string from a cell, or None if no formula is set."""
        wb = openpyxl.load_workbook(str(path))
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found")
        ws = wb[sheet_name]
        value = ws[cell_ref].value
        if isinstance(value, str) and value.startswith("="):
            return value
        return None

    @staticmethod
    def get_sheet_info(path: Path, sheet_name: str) -> GetSheetInfoResponse:
        wb = openpyxl.load_workbook(str(path), data_only=True)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found")
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]] if ws.max_row > 0 else []
        return GetSheetInfoResponse(
            sheet_name=sheet_name,
            row_count=ws.max_row or 0,
            col_count=ws.max_column or 0,
            headers=headers,
        )

    @staticmethod
    def list_sheets(path: Path) -> SpreadsheetInfo:
        wb = openpyxl.load_workbook(str(path), data_only=True)
        sheet_infos = []
        for name in wb.sheetnames:
            ws = wb[name]
            sheet_infos.append(
                SheetInfo(name=name, row_count=ws.max_row or 0, col_count=ws.max_column or 0)
            )
        return SpreadsheetInfo(file_path=str(path), sheets=sheet_infos)


def get_engine(path: Path) -> OdsEngine | XlsxEngine:
    ftype = _get_file_type(path)
    if ftype == "ods":
        return OdsEngine()
    return XlsxEngine()
